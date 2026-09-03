#!/usr/bin/env python3
"""
informatica_lineage_extractor.py
=================================

Search an Informatica PowerCenter repository export (XML, or the
tag/attributes/children-style JSON produced by converting that XML) for every
place a given field/attribute name is used, and write the results to an
Excel workbook.

------------------------------------------------------------------------
INPUT
------------------------------------------------------------------------
Accepts either:
  * The native PowerCenter XML export (<POWERMART>...</POWERMART>), or
  * A JSON file that is the direct tag/attributes/children conversion of
    that XML, shaped like:

        {
          "xml_declaration": "...",
          "doctype": "...",
          "root": {
              "tag": "POWERMART",
              "attributes": {...},
              "children": [ {"tag": "...", "attributes": {...}, "children":[...]}, ... ]
          }
        }

The file type is auto-detected from the extension (.xml / .json) and,
if that fails, from the file's first non-whitespace character.

------------------------------------------------------------------------
WHAT IT SEARCHES
------------------------------------------------------------------------
For the attribute (port/field) name you give it, the script walks every
MAPPING (and MAPPLET) in the repository and inspects each TRANSFORMATION,
dispatching to a handler for its TYPE. Every handler looks in the two
places business logic actually lives in a PowerCenter export - the port
(TRANSFORMFIELD, e.g. an EXPRESSION) and the transformation-level
properties (TABLEATTRIBUTE, e.g. a Filter/Join/Lookup Condition) - but
which one it looks at, and how it labels what it finds, depends on the
transformation type:

  Expression                -> the port's own expression (unchanged style:
                                "field" for pass-through, "field=expr" for
                                a real calculation)
  Aggregator                -> Aggregator Expression (same pass-through/
                                calculation test as Expression, but always
                                labelled "Aggregator Expression: ...")
  Filter                    -> Filter Condition
  Joiner                    -> Join Condition
  Lookup Procedure          -> whichever of Lookup Condition / Lookup Sql
                                Override / Lookup Table Name references it
                                (only the ones that actually match are shown)
  Router                    -> Group Name + that group's Filter Condition
  Sorter                    -> Sort Direction (if it's a sort key) +
                                Transformation Scope
  Source Qualifier          -> Business_Logic = Sql Query;
                                Additional_Information = User Defined Join,
                                Source Filter, Associated Source Definitions
  Stored Procedure          -> Stored Procedure Name + whichever property
                                (Connection Information, Stored Procedure
                                Type, Execution Order, etc.) references it
  Transaction Control       -> Transaction Control Condition
  Update Strategy           -> Update Strategy Expression
  Rank                      -> Top/Bottom + Number of Ranks
  anything else             -> generic fallback: port expression scan +
                                a broad scan of every TABLEATTRIBUTE value

Each mapping is then connected back to the Workflow(s)/Session(s) that run
it (WORKFLOW -> SESSION -> MAPPINGNAME -> MAPPING), and to the upstream /
downstream ports via the mapping's CONNECTOR elements, so every occurrence
row also carries its predecessor and successor port.

Mapplets referenced from a mapping (INSTANCE TYPE="MAPPLET") are expanded
and searched too; their internal predecessor/successor linkage is resolved
within the mapplet's own connector graph (the link from the mapplet's own
boundary ports back out into the parent mapping is not attempted - see the
"KNOWN LIMITATIONS" note in the README section printed with --help).

------------------------------------------------------------------------
"Direct Pass through" vs "Involves Derivation"
------------------------------------------------------------------------
Every occurrence is classified as one or the other:
  - "Direct Pass through": the attribute is simply carried through
    unchanged (a bare port, an expression that's just a single identifier,
    a pass-through group in a Router, a non-sort-key port in a Sorter, a
    lookup port not referenced by any lookup condition, etc.)
  - "Involves Derivation": the attribute actually participates in some
    logic - a function/operator/condition in an expression, a Filter/Join/
    Lookup/Router/Transaction-Control condition, a Sql Query/Source Filter/
    User Defined Join, being an active sort or rank key, etc.

------------------------------------------------------------------------
OUTPUT
------------------------------------------------------------------------
An .xlsx workbook with:
  Tab 1 "Occurrences"       - one row per occurrence, exactly the columns
      below: Workflow_name, Session_name, Mapping_name, Transformation_name,
      Transformation_type, Transformation_Field, Business_Logic,
      Category, Predecessor_Input, Successor_Output, Additional_Information
      (Additional_Information is populated for Source Qualifier rows only;
      blank everywhere else.)
  Tab 2 "Reporting_Lineage" - only built when --target-mapping and
      --target-table are given (or supplied when prompted). Contains one row
      for EVERY occurrence of the searched attribute found anywhere in the
      repository - the same set of occurrences as Tab 1 - each traced
      forward through the CONNECTOR graph of its OWN mapping (crossing
      into/out of mapplets at their Input/Output boundary, and following a
      field as it evolves into a differently-named derived port) until it
      either reaches a TARGET instance or runs out of connectors. Columns:
      Mapping_name, Transformation_name, Transformation_type,
      Transformation_Field, Lineage (e.g.
      "EXP1.IN_X -> EXP1.OUT_Y -> TGT1.OUT_Y"), Status
      ("Involved Reporting Lineage" only for occurrences that are inside the
      given --target-mapping AND whose traced lineage reaches the given
      --target-table there; every other occurrence is still listed, marked
      "Not_Involved").
  Tab 3 "Summary"           - quick counts, useful for a large result set.

------------------------------------------------------------------------
USAGE
------------------------------------------------------------------------
    python informatica_lineage_extractor.py --input export.xml --attribute ISSUER
    python informatica_lineage_extractor.py --input export.json --attribute ISSUER -o result.xlsx

    # case-sensitive matching (default is case-insensitive):
    python informatica_lineage_extractor.py -i export.xml -a ISSUER --case-sensitive

    # skip expanding mapplets (faster on huge repos if you don't need them):
    python informatica_lineage_extractor.py -i export.xml -a ISSUER --no-mapplets

    # also build the Reporting_Lineage tab (Tab 2):
    python informatica_lineage_extractor.py -i export.xml -a ISSUER \
        --target-mapping m_REPORTING_LOAD --target-table TGT_REPORT_TABLE

    # skip the Reporting_Lineage tab / its prompt entirely:
    python informatica_lineage_extractor.py -i export.xml -a ISSUER --no-lineage-tab
"""

import argparse
import json
import re
import sys
from collections import defaultdict

try:
    import xml.etree.ElementTree as ET
except ImportError:  # pragma: no cover
    ET = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("This script needs openpyxl. Install it with: pip install openpyxl")


BARE_IDENTIFIER_RE = re.compile(r'^[A-Za-z_][A-Za-z0-9_]*$')

DIRECT = "Direct Pass through"
INDIRECT = "Involves Derivation"


# ---------------------------------------------------------------------------
# 1. Loading: XML or JSON -> a common {tag, attributes, children} tree
# ---------------------------------------------------------------------------

def xml_element_to_dict(elem):
    """Convert an ElementTree element into the tag/attributes/children shape."""
    node = {"tag": elem.tag, "attributes": dict(elem.attrib)}
    kids = [xml_element_to_dict(c) for c in elem]
    if kids:
        node["children"] = kids
    return node


def load_tree(path):
    """Load the input file (XML or JSON) and return the root node dict."""
    is_json = path.lower().endswith(".json")
    is_xml = path.lower().endswith(".xml")

    if not (is_json or is_xml):
        # sniff the file
        with open(path, "rb") as f:
            head = f.read(2048).lstrip()
        is_xml = head.startswith(b"<")
        is_json = head.startswith(b"{") or head.startswith(b"[")

    if is_xml:
        if ET is None:
            sys.exit("xml.etree.ElementTree is unavailable in this environment.")
        print(f"Parsing XML: {path} ...")
        tree = ET.parse(path)
        return xml_element_to_dict(tree.getroot())

    print(f"Loading JSON: {path} ...")
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    if isinstance(data, dict) and "root" in data and "tag" in data.get("root", {}):
        return data["root"]
    if isinstance(data, dict) and "tag" in data:
        return data
    sys.exit("Could not find a {'tag':..., 'attributes':..., 'children':...} "
              "root object in the JSON file.")


# ---------------------------------------------------------------------------
# 2. Small tree helpers
# ---------------------------------------------------------------------------

def children(node, tag=None):
    """Direct children of node, optionally filtered by tag."""
    for c in node.get("children") or []:
        if isinstance(c, dict) and "tag" in c and (tag is None or c["tag"] == tag):
            yield c


def find_all(node, tag):
    """All descendants of node (at any depth) matching tag."""
    for c in node.get("children") or []:
        if not isinstance(c, dict) or "tag" not in c:
            continue
        if c["tag"] == tag:
            yield c
        yield from find_all(c, tag)


def attr(node, name, default=""):
    return node.get("attributes", {}).get(name, default)


def contains_token(text, attribute, case_sensitive):
    """True if `attribute` appears as a whole word/token inside `text`."""
    if not text:
        return False
    flags = 0 if case_sensitive else re.IGNORECASE
    pattern = r'(?<![A-Za-z0-9_])' + re.escape(attribute) + r'(?![A-Za-z0-9_])'
    return re.search(pattern, text, flags) is not None


def names_equal(a, b, case_sensitive):
    if a is None or b is None:
        return False
    return a == b if case_sensitive else a.lower() == b.lower()


# ---------------------------------------------------------------------------
# 3. Repository indexing
# ---------------------------------------------------------------------------

class Repository:
    """Indexes folders / mappings / mapplets / workflows once, up front."""

    def __init__(self, root):
        self.root = root
        self.folders = list(find_all(root, "FOLDER"))

        # global fallbacks, used only if a session's mapping isn't found
        # in its own folder (e.g. shortcut to a shared folder)
        self.global_mappings = {}
        self.global_mapplets = {}

        self.folder_mappings = {}   # folder_name -> {mapping_name: node}
        self.folder_mapplets = {}   # folder_name -> {mapplet_name: node}

        # reusable transformations defined directly under FOLDER (referenced
        # from a mapping/mapplet by shortcut, not embedded inline)
        self.global_transformations = {}
        self.folder_transformations = {}

        for folder in self.folders:
            fname = attr(folder, "NAME")
            mmap = {}
            for m in children(folder, "MAPPING"):
                mmap[attr(m, "NAME")] = m
                self.global_mappings.setdefault(attr(m, "NAME"), m)
            self.folder_mappings[fname] = mmap

            plmap = {}
            for m in children(folder, "MAPPLET"):
                plmap[attr(m, "NAME")] = m
                self.global_mapplets.setdefault(attr(m, "NAME"), m)
            self.folder_mapplets[fname] = plmap

            tmap = {}
            for t in children(folder, "TRANSFORMATION"):
                tmap[attr(t, "NAME")] = t
                self.global_transformations.setdefault(attr(t, "NAME"), t)
            self.folder_transformations[fname] = tmap

    def get_mapping(self, folder_name, mapping_name):
        node = self.folder_mappings.get(folder_name, {}).get(mapping_name)
        return node or self.global_mappings.get(mapping_name)

    def get_mapplet(self, folder_name, mapplet_name):
        node = self.folder_mapplets.get(folder_name, {}).get(mapplet_name)
        return node or self.global_mapplets.get(mapplet_name)

    def get_reusable_transformation(self, folder_name, t_name):
        node = self.folder_transformations.get(folder_name, {}).get(t_name)
        return node or self.global_transformations.get(t_name)


# ---------------------------------------------------------------------------
# 4. Connector graph + per-mapping/mapplet search
# ---------------------------------------------------------------------------

def build_connector_index(container):
    """
    Build predecessor/successor lookups from a MAPPING's or MAPPLET's
    CONNECTOR children.
      incoming[(instance, field)] -> list of "from_instance.from_field"
      outgoing[(instance, field)] -> list of "to_instance.to_field"
    """
    incoming = defaultdict(list)
    outgoing = defaultdict(list)
    for c in children(container, "CONNECTOR"):
        a = c.get("attributes", {})
        from_i, from_f = a.get("FROMINSTANCE", ""), a.get("FROMFIELD", "")
        to_i, to_f = a.get("TOINSTANCE", ""), a.get("TOFIELD", "")
        incoming[(to_i, to_f)].append(f"{from_i}.{from_f}")
        outgoing[(from_i, from_f)].append(f"{to_i}.{to_f}")
    return incoming, outgoing


def build_instance_map(container):
    """TRANSFORMATION_NAME -> [INSTANCE node, ...] for a MAPPING/MAPPLET."""
    m = defaultdict(list)
    for inst in children(container, "INSTANCE"):
        t_name = attr(inst, "TRANSFORMATION_NAME")
        m[t_name].append(inst)
    return m


def get_tableattr(t, target_name):
    """Case-insensitive exact lookup of a TABLEATTRIBUTE VALUE by NAME."""
    target = target_name.strip().lower()
    for ta in children(t, "TABLEATTRIBUTE"):
        if attr(ta, "NAME", "").strip().lower() == target:
            return attr(ta, "VALUE", "")
    return None


def get_tableattr_fuzzy(t, name_variants):
    """Like get_tableattr, but tries several possible attribute names
    (schemas/versions of PowerCenter spell some of these differently)."""
    variants = {n.strip().lower() for n in name_variants}
    for ta in children(t, "TABLEATTRIBUTE"):
        if attr(ta, "NAME", "").strip().lower() in variants:
            return attr(ta, "VALUE", "")
    return None


def scan_ports(t, attribute, case_sensitive):
    """Every TRANSFORMFIELD on t whose NAME or EXPRESSION references the
    attribute. Each hit dict carries enough to classify it further."""
    hits = []
    for f in children(t, "TRANSFORMFIELD"):
        f_name = attr(f, "NAME")
        expr = attr(f, "EXPRESSION", "")
        name_hit = names_equal(f_name, attribute, case_sensitive)
        expr_hit = contains_token(expr, attribute, case_sensitive) if expr else False
        if name_hit or expr_hit:
            hits.append({"node": f, "name": f_name, "expression": expr,
                         "name_hit": name_hit, "expr_hit": expr_hit})
    return hits


def classify_expression_hit(field_name, expression, name_hit):
    """
    Expression-style classification (used for Expression & Aggregator ports,
    and as the generic fallback for any other type's ports).
    Returns (category, business_logic).
    """
    if not expression:
        return DIRECT, field_name

    is_bare = bool(BARE_IDENTIFIER_RE.match(expression.strip()))

    if name_hit and expression.strip() == field_name:
        return DIRECT, field_name
    if is_bare:
        return DIRECT, field_name if expression.strip() == field_name else f"{field_name}={expression}"
    return INDIRECT, f"{field_name}={expression}"


def field_lineage(incoming, outgoing, instance_names, field_name):
    preds, succs = [], []
    for inst in instance_names:
        preds += incoming.get((inst, field_name), [])
        succs += outgoing.get((inst, field_name), [])
    return "; ".join(preds), "; ".join(succs)


def make_row(t_name, t_type, field, business_logic, category,
             predecessor="", successor="", additional_info=""):
    return {
        "transformation_name": t_name, "transformation_type": t_type,
        "field": field, "business_logic": business_logic, "category": category,
        "predecessor": predecessor, "successor": successor,
        "additional_info": additional_info,
    }


# ---- per-transformation-type handlers --------------------------------
#
# Each handler receives:
#   t              - the TRANSFORMATION node
#   attribute, case_sensitive
#   incoming, outgoing  - connector lineage dicts for the whole container
#   instance_names  - the INSTANCE name(s) this transformation is placed as
#   instance_nodes  - the INSTANCE node(s) themselves (for e.g. associated sources)
# and returns a list of match dicts (via make_row), WITHOUT the label_prefix
# applied to transformation_name (the caller adds that).

def _handle_expression_style(t, attribute, case_sensitive, incoming, outgoing,
                              instance_names, instance_nodes):
    """Expression & Aggregator: business logic = the port's own expression."""
    rows = []
    for f in children(t, "TRANSFORMFIELD"):
        f_name = attr(f, "NAME")
        expr = attr(f, "EXPRESSION", "")
        name_hit = names_equal(f_name, attribute, case_sensitive)
        expr_hit = contains_token(expr, attribute, case_sensitive) if expr else False
        if not (name_hit or expr_hit):
            continue
        category, logic = classify_expression_hit(f_name, expr, name_hit)
        pred, succ = field_lineage(incoming, outgoing, instance_names, f_name)
        rows.append(make_row(None, None, f_name, logic, category, pred, succ))
    return rows


def _handle_single_condition(t, attribute, case_sensitive, incoming, outgoing,
                              instance_names, instance_nodes,
                              tableattr_names, label):
    """Filter / Joiner / Update Strategy / Transaction Control: one
    transformation-level condition, plus plain pass-through ports."""
    value = get_tableattr_fuzzy(t, tableattr_names)
    port_hits = scan_ports(t, attribute, case_sensitive)
    cond_hit = contains_token(value, attribute, case_sensitive) if value else False

    if not port_hits and not cond_hit:
        return []

    category = INDIRECT if cond_hit else DIRECT
    field_names = sorted({h["name"] for h in port_hits}) or [attribute]
    preds, succs = [], []
    for fn in field_names:
        p, s = field_lineage(incoming, outgoing, instance_names, fn)
        if p:
            preds.append(p)
        if s:
            succs.append(s)
    business_logic = f"{label}: {value}" if value is not None else f"{label}: (not set)"
    return [make_row(None, None, "; ".join(field_names), business_logic, category,
                      "; ".join(preds), "; ".join(succs))]


def _handle_aggregator(t, attribute, case_sensitive, incoming, outgoing,
                        instance_names, instance_nodes):
    """Aggregator: same Direct/Derivation test as Expression (does the port's
    own expression actually do anything), but the Business Logic value is
    always labelled 'Aggregator Expression: ...' rather than left bare."""
    rows = []
    for f in children(t, "TRANSFORMFIELD"):
        f_name = attr(f, "NAME")
        expr = attr(f, "EXPRESSION", "")
        name_hit = names_equal(f_name, attribute, case_sensitive)
        expr_hit = contains_token(expr, attribute, case_sensitive) if expr else False
        if not (name_hit or expr_hit):
            continue
        category, _ = classify_expression_hit(f_name, expr, name_hit)
        value = expr if expr else f_name
        logic = f"Aggregator Expression: {value}"
        pred, succ = field_lineage(incoming, outgoing, instance_names, f_name)
        rows.append(make_row(None, None, f_name, logic, category, pred, succ))
    return rows


def _handle_lookup(t, attribute, case_sensitive, incoming, outgoing,
                    instance_names, instance_nodes):
    """Lookup Procedure: up to 3 possible business-logic sources, shown only
    where they actually reference the attribute, plus plain pass-through
    lookup ports."""
    rows = []
    checks = [
        (("Lookup condition", "Lookup Condition"), "Lookup Condition"),
        (("Lookup Sql Override",), "Lookup Sql Override"),
        (("Lookup table name", "Lookup Table Name"), "Lookup Table Name"),
    ]
    port_hits = scan_ports(t, attribute, case_sensitive)
    field_names = sorted({h["name"] for h in port_hits})

    any_condition_hit = False
    for variants, label in checks:
        value = get_tableattr_fuzzy(t, variants)
        if value and contains_token(value, attribute, case_sensitive):
            any_condition_hit = True
            preds, succs = [], []
            for fn in field_names:
                p, s = field_lineage(incoming, outgoing, instance_names, fn)
                if p:
                    preds.append(p)
                if s:
                    succs.append(s)
            rows.append(make_row(
                None, None, "; ".join(field_names) or attribute,
                f"{label}: {value}", INDIRECT, "; ".join(preds), "; ".join(succs)))

    if port_hits and not any_condition_hit:
        # plain pass-through lookup port(s), unrelated to the 3 conditions above
        for h in port_hits:
            pred, succ = field_lineage(incoming, outgoing, instance_names, h["name"])
            rows.append(make_row(None, None, h["name"], h["name"], DIRECT, pred, succ))

    return rows


def _handle_router(t, attribute, case_sensitive, incoming, outgoing,
                    instance_names, instance_nodes):
    """Router: one row per GROUP where the attribute shows up, either as a
    pass-through port in that group or inside the group's filter condition."""
    rows = []
    for g in children(t, "GROUP"):
        g_name = attr(g, "NAME")
        g_expr = attr(g, "EXPRESSION", "")
        name_hit = any(
            attr(f, "GROUP") == g_name and names_equal(attr(f, "NAME"), attribute, case_sensitive)
            for f in children(t, "TRANSFORMFIELD"))
        expr_hit = contains_token(g_expr, attribute, case_sensitive) if g_expr else False
        if not (name_hit or expr_hit):
            continue
        category = INDIRECT if expr_hit else DIRECT
        logic = f"Group Name: {g_name}"
        if g_expr:
            logic += f"; Group Filter Condition: {g_expr}"
        pred, succ = field_lineage(incoming, outgoing, instance_names, attribute)
        rows.append(make_row(None, None, attribute, logic, category, pred, succ))
    return rows


def _handle_sorter(t, attribute, case_sensitive, incoming, outgoing,
                    instance_names, instance_nodes):
    """Sorter: business logic = sort direction (per port) + transformation
    scope; category depends on whether the attribute is an actual sort key."""
    rows = []
    scope = get_tableattr(t, "Transformation Scope") or "N/A"
    for f in children(t, "TRANSFORMFIELD"):
        f_name = attr(f, "NAME")
        if not names_equal(f_name, attribute, case_sensitive):
            continue
        is_sort_key = attr(f, "ISSORTKEY", "NO").upper() == "YES"
        sort_dir = attr(f, "SORTDIRECTION", "N/A") if is_sort_key else "N/A (not a sort key)"
        category = INDIRECT if is_sort_key else DIRECT
        logic = f"Sort Direction: {sort_dir}; Transformation Scope: {scope}"
        pred, succ = field_lineage(incoming, outgoing, instance_names, f_name)
        rows.append(make_row(None, None, f_name, logic, category, pred, succ))
    return rows


def _handle_source_qualifier(t, attribute, case_sensitive, incoming, outgoing,
                              instance_names, instance_nodes):
    """Source Qualifier: Business Logic = Sql Query; Additional Information =
    User Defined Join / Source Filter / Associated Source Definitions."""
    sql = get_tableattr(t, "Sql Query") or ""
    udj = get_tableattr(t, "User Defined Join") or ""
    srcfilter = get_tableattr(t, "Source Filter") or ""

    assoc_names = []
    for inst in instance_nodes:
        for a in children(inst, "ASSOCIATED_SOURCE_INSTANCE"):
            n = attr(a, "NAME")
            if n:
                assoc_names.append(n)

    port_hits = scan_ports(t, attribute, case_sensitive)
    hit_in_logic = any(
        contains_token(v, attribute, case_sensitive) for v in (sql, udj, srcfilter))

    if not port_hits and not hit_in_logic:
        return []

    category = INDIRECT if hit_in_logic else DIRECT
    field_names = sorted({h["name"] for h in port_hits}) or [attribute]
    preds, succs = [], []
    for fn in field_names:
        p, s = field_lineage(incoming, outgoing, instance_names, fn)
        if p:
            preds.append(p)
        if s:
            succs.append(s)

    business_logic = f"SQL Query: {sql}" if sql else "SQL Query: (none - default generated query)"
    additional_info = (
        f"User Defined Join: {udj or '(none)'}; "
        f"Source Filter: {srcfilter or '(none)'}; "
        f"Associated Source Definitions: {', '.join(assoc_names) or '(none)'}")

    return [make_row(None, None, "; ".join(field_names), business_logic, category,
                      "; ".join(preds), "; ".join(succs), additional_info)]


def _handle_stored_procedure(t, attribute, case_sensitive, incoming, outgoing,
                              instance_names, instance_nodes):
    """Stored Procedure: business logic = SP name + whichever property
    (Connection Information / Stored Procedure Type / Execution Order /
    any other property) references the attribute."""
    rows = []
    sp_name = get_tableattr_fuzzy(t, ["Stored Procedure Name"]) or attr(t, "NAME")
    port_hits = scan_ports(t, attribute, case_sensitive)
    field_names = sorted({h["name"] for h in port_hits})

    any_prop_hit = False
    for ta in children(t, "TABLEATTRIBUTE"):
        ta_name = attr(ta, "NAME")
        ta_val = attr(ta, "VALUE", "")
        if ta_name.strip().lower() == "stored procedure name":
            continue
        if ta_val and contains_token(ta_val, attribute, case_sensitive):
            any_prop_hit = True
            preds, succs = [], []
            for fn in field_names:
                p, s = field_lineage(incoming, outgoing, instance_names, fn)
                if p:
                    preds.append(p)
                if s:
                    succs.append(s)
            rows.append(make_row(
                None, None, "; ".join(field_names) or attribute,
                f"Stored Procedure Name: {sp_name}; {ta_name}: {ta_val}", INDIRECT,
                "; ".join(preds), "; ".join(succs)))

    if port_hits and not any_prop_hit:
        for h in port_hits:
            pred, succ = field_lineage(incoming, outgoing, instance_names, h["name"])
            rows.append(make_row(
                None, None, h["name"], f"Stored Procedure Name: {sp_name}; Port: {h['name']}",
                DIRECT, pred, succ))

    return rows


def _handle_rank(t, attribute, case_sensitive, incoming, outgoing,
                  instance_names, instance_nodes):
    """Rank: business logic = Top/Bottom + Number of Ranks; category depends
    on whether the attribute is the actual rank port."""
    rows = []
    topbottom = get_tableattr_fuzzy(t, ["Top/Bottom", "Rank Top/Bottom"]) or "N/A"
    num_ranks = get_tableattr_fuzzy(t, ["Number Of Ranks", "Number of Ranks"]) or "N/A"
    for f in children(t, "TRANSFORMFIELD"):
        f_name = attr(f, "NAME")
        if not names_equal(f_name, attribute, case_sensitive):
            continue
        is_rank_port = attr(f, "ISRANKPORT", attr(f, "RANKPORT", "NO")).upper() == "YES"
        category = INDIRECT if is_rank_port else DIRECT
        logic = f"Top/Bottom: {topbottom}; Number of Ranks: {num_ranks}"
        pred, succ = field_lineage(incoming, outgoing, instance_names, f_name)
        rows.append(make_row(None, None, f_name, logic, category, pred, succ))
    return rows


def _handle_fallback(t, attribute, case_sensitive, incoming, outgoing,
                      instance_names, instance_nodes):
    """Any transformation TYPE without a dedicated handler above: generic
    expression-style port scan + a broad scan of every TABLEATTRIBUTE."""
    rows = _handle_expression_style(t, attribute, case_sensitive, incoming, outgoing,
                                     instance_names, instance_nodes)
    for ta in children(t, "TABLEATTRIBUTE"):
        ta_name = attr(ta, "NAME")
        ta_val = attr(ta, "VALUE", "")
        if not ta_val:
            continue
        if contains_token(ta_val, attribute, case_sensitive):
            rows.append(make_row(None, None, f"[{ta_name}]", f"{ta_name}: {ta_val}", INDIRECT))
    return rows


TYPE_HANDLERS = {
    "Expression": _handle_expression_style,
    "Aggregator": _handle_aggregator,
    "Filter": lambda *a: _handle_single_condition(*a, ["Filter Condition"], "Filter Condition"),
    "Joiner": lambda *a: _handle_single_condition(*a, ["Join Condition"], "Join Condition"),
    "Update Strategy": lambda *a: _handle_single_condition(
        *a, ["Update Strategy Expression"], "Update Strategy Expression"),
    "Transaction Control": lambda *a: _handle_single_condition(
        *a, ["Transaction Control Condition", "Transaction Control Expression"],
        "Transaction Control Condition"),
    "Lookup Procedure": _handle_lookup,
    "Router": _handle_router,
    "Sorter": _handle_sorter,
    "Source Qualifier": _handle_source_qualifier,
    "Stored Procedure": _handle_stored_procedure,
    "Rank": _handle_rank,
}


def search_container(container, attribute, case_sensitive, label_prefix=""):
    """
    Search every TRANSFORMATION inside a MAPPING or MAPPLET node, dispatching
    to the transformation-type-specific handler (falling back to a generic
    one for any TYPE not explicitly covered).
    Returns a list of match dicts (mapping/session/workflow-agnostic):
        transformation_name, transformation_type, field, business_logic,
        category, predecessor, successor, additional_info
    """
    incoming, outgoing = build_connector_index(container)
    inst_map = build_instance_map(container)
    results = []

    for t in children(container, "TRANSFORMATION"):
        t_name = attr(t, "NAME")
        t_type = attr(t, "TYPE")
        instance_nodes = inst_map.get(t_name) or []
        instance_names = [attr(n, "NAME") for n in instance_nodes] or [t_name]

        handler = TYPE_HANDLERS.get(t_type, _handle_fallback)
        matches = handler(t, attribute, case_sensitive, incoming, outgoing,
                           instance_names, instance_nodes)

        for m in matches:
            m["transformation_name"] = f"{label_prefix}{t_name}"
            m["transformation_type"] = t_type
            results.append(m)

    return results


def search_mapping_with_mapplets(repo, folder_name, mapping, attribute,
                                  case_sensitive, expand_mapplets, cache):
    """
    Search a mapping, and (optionally) expand any MAPPLET instances it uses,
    searching inside those too. Cached per mapping node id.
    """
    key = id(mapping)
    if key in cache:
        return cache[key]

    results = search_container(mapping, attribute, case_sensitive)

    if expand_mapplets:
        for inst in children(mapping, "INSTANCE"):
            if attr(inst, "TYPE") != "MAPPLET":
                continue
            mapplet_name = attr(inst, "TRANSFORMATION_NAME")
            mapplet = repo.get_mapplet(folder_name, mapplet_name)
            if mapplet is None:
                continue
            inner = search_container(
                mapplet, attribute, case_sensitive,
                label_prefix=f"{mapplet_name}::")
            results.extend(inner)

    cache[key] = results
    return results


# ---------------------------------------------------------------------------
# 5. Reporting-lineage tracing (Tab 2: "Reporting_Lineage")
# ---------------------------------------------------------------------------
#
# Given one occurrence found above (an instance.field where the searched
# attribute either sits directly on a port or drives some derivation), this
# walks the CONNECTOR graph forward - hopping instance by instance, and
# jumping into/out of MAPPLET instances at their Input/Output boundary -
# until it either:
#   * lands on a TARGET instance (a sink)              -> status "sink"
#   * runs out of outgoing connectors                  -> status "deadend"
#   * loops back on an already-visited (instance,field) -> status "deadend"
# Whenever the field lands on an INPUT-only port of a transformation, it
# also looks at every OUTPUT port on that same transformation whose
# EXPRESSION references the field by name, and continues down each one -
# this is how a field that "evolves" into a differently named derived
# port (e.g. an Expression computing OUT_TOTAL from IN_AMOUNT) is followed.
#
# NOTE / assumption: like the rest of this script, an INSTANCE is assumed to
# share its NAME with the TRANSFORMATION/SOURCE/TARGET it points to whenever
# an explicit INSTANCE node can't be resolved (the same fallback already
# used elsewhere for predecessor/successor lookups). This holds for the
# overwhelming majority of real PowerCenter exports (one instance per
# embedded transformation, named the same).

def build_connector_index_tuples(container):
    """Like build_connector_index, but keeps (instance, field) as a real
    tuple instead of a joined string, since the lineage tracer needs to
    keep hopping instance-by-instance."""
    outgoing = defaultdict(list)
    for c in children(container, "CONNECTOR"):
        a = c.get("attributes", {})
        from_i, from_f = a.get("FROMINSTANCE", ""), a.get("FROMFIELD", "")
        to_i, to_f = a.get("TOINSTANCE", ""), a.get("TOFIELD", "")
        outgoing[(from_i, from_f)].append((to_i, to_f))
    return outgoing


def make_lineage_ctx(container, parent_ctx=None, parent_instance_name=None):
    """Everything the tracer needs about one MAPPING/MAPPLET container.
    parent_ctx/parent_instance_name are set only when this container is a
    MAPPLET being entered from a MAPPLET instance in parent_ctx - they're
    how the tracer bridges back out once it reaches the mapplet's Output
    boundary transformation."""
    return {
        "outgoing": build_connector_index_tuples(container),
        "inst_by_name": {attr(i, "NAME"): i for i in children(container, "INSTANCE")},
        "t_by_name": {attr(t, "NAME"): t for t in children(container, "TRANSFORMATION")},
        "parent_ctx": parent_ctx,
        "parent_instance_name": parent_instance_name,
    }


def _lineage_instance_type_and_target(ctx, instance_name):
    """(INSTANCE TYPE upper-cased, the name it points to) or (None, None)."""
    inst = ctx["inst_by_name"].get(instance_name)
    if inst is None:
        return None, None
    tname = attr(inst, "TRANSFORMATION_NAME") or attr(inst, "NAME")
    return attr(inst, "TYPE", "").upper(), tname


def _lineage_transformation_node(ctx, folder_name, repo, tname):
    node = ctx["t_by_name"].get(tname)
    if node is None:
        node = repo.get_reusable_transformation(folder_name, tname)
    return node


def _lineage_find_boundary_instance(ctx, keyword):
    """Find the INSTANCE inside a mapplet whose underlying TRANSFORMATION
    TYPE contains 'input' or 'output' (the mapplet's boundary ports)."""
    for inst_name, inst in ctx["inst_by_name"].items():
        if attr(inst, "TYPE", "").upper() != "TRANSFORMATION":
            continue
        tname = attr(inst, "TRANSFORMATION_NAME") or attr(inst, "NAME")
        t_node = ctx["t_by_name"].get(tname)
        if t_node is None:
            continue
        if keyword in attr(t_node, "TYPE", "").lower():
            return inst_name
    return None


def _lineage_resolve_outputs(t_node, field_name):
    """Given the field as it arrives on this transformation, return the
    output-port field name(s) that carry it onward: itself if it's already
    an output (or pass-through input/output) port, or every output port
    whose EXPRESSION references it, if it's an input-only port."""
    f = None
    for tf in children(t_node, "TRANSFORMFIELD"):
        if names_equal(attr(tf, "NAME"), field_name, case_sensitive=False):
            f = tf
            break
    if f is None:
        return []
    porttype = attr(f, "PORTTYPE", "").upper()
    if "OUTPUT" in porttype:
        return [attr(f, "NAME")]
    outs = []
    for g in children(t_node, "TRANSFORMFIELD"):
        if "OUTPUT" not in attr(g, "PORTTYPE", "").upper():
            continue
        g_name = attr(g, "NAME")
        if names_equal(g_name, field_name, case_sensitive=False):
            continue
        g_expr = attr(g, "EXPRESSION", "")
        if g_expr and contains_token(g_expr, field_name, case_sensitive=False):
            outs.append(g_name)
    return outs


def _lineage_is_unconnected_capable(t_node):
    """Lookup Procedure and Stored Procedure transformations can be invoked
    "unconnected" in PowerCenter - called from inside another port's
    EXPRESSION via ':LKP.<name>(...)' or ':SP.<name>(...)' instead of being
    wired through CONNECTOR elements like every other transformation. An
    unconnected instance therefore has no incoming/outgoing CONNECTOR at
    all, so the only way to keep tracing its lineage forward is to find
    whichever port(s) actually call it."""
    ttype = attr(t_node, "TYPE", "").upper()
    return "LOOKUP" in ttype or "STORED PROCEDURE" in ttype


_unconnected_call_re_cache = {}


def _lineage_find_unconnected_callers(ctx, tname):
    """Every (instance_name, output_field_name) inside this same container
    (mapping or mapplet) whose EXPRESSION invokes `tname` as an unconnected
    Lookup/Stored Procedure - i.e. contains ':LKP.tname(' or ':SP.tname('
    (case-insensitive, ':' prefix and '(' required so a same-named ordinary
    port reference doesn't false-match)."""
    pattern = _unconnected_call_re_cache.get(tname)
    if pattern is None:
        pattern = re.compile(r":(?:LKP|SP)\." + re.escape(tname) + r"\s*\(", re.IGNORECASE)
        _unconnected_call_re_cache[tname] = pattern

    hits = []
    for inst_name, inst in ctx["inst_by_name"].items():
        if attr(inst, "TYPE", "").upper() != "TRANSFORMATION":
            continue
        caller_tname = attr(inst, "TRANSFORMATION_NAME") or attr(inst, "NAME")
        if names_equal(caller_tname, tname, False):
            continue  # a lookup/SP calling itself isn't meaningful here
        t_node = ctx["t_by_name"].get(caller_tname)
        if t_node is None:
            continue
        for f in children(t_node, "TRANSFORMFIELD"):
            if "OUTPUT" not in attr(f, "PORTTYPE", "").upper():
                continue
            expr = attr(f, "EXPRESSION", "")
            if expr and pattern.search(expr):
                hits.append((inst_name, attr(f, "NAME")))
    return hits


def _lineage_unconnected_fallback(ctx, folder_name, repo, t_node, tname,
                                   visited, sub_path, depth, max_depth):
    """When a Lookup/Stored-Procedure instance has no outgoing CONNECTOR
    (a "connected" dead end), check whether it's actually invoked
    unconnected elsewhere in the same mapping/mapplet and, if so, continue
    the trace into every call site instead of reporting a dead end. Returns
    None (meaning: no unconnected call site found, caller should fall back
    to its normal dead-end handling) or a list of trace results."""
    if t_node is None or not _lineage_is_unconnected_capable(t_node):
        return None
    callers = _lineage_find_unconnected_callers(ctx, tname)
    if not callers:
        return None
    results = []
    for ni, nf in callers:
        results += trace_lineage(ctx, folder_name, repo, ni, nf, visited, sub_path, depth + 1, max_depth)
    return results


def trace_lineage(ctx, folder_name, repo, instance_name, field_name,
                   visited, path, depth=0, max_depth=250):
    """Forward-trace one (instance, field) to every sink/dead-end it reaches.
    Returns a list of {"path": [...], "status": "sink"|"deadend", ...}."""
    node_label = f"{instance_name}.{field_name}"
    if depth > max_depth:
        return [{"path": path + [node_label], "status": "deadend"}]

    key = (id(ctx), instance_name, field_name)
    if key in visited:
        return [{"path": path + [node_label], "status": "deadend"}]
    visited = visited | {key}
    path = path + [node_label]

    itype, tname = _lineage_instance_type_and_target(ctx, instance_name)
    if itype is None:
        return [{"path": path, "status": "deadend"}]

    if itype == "TARGET":
        return [{"path": path, "status": "sink", "sink_tname": tname, "sink_instance": instance_name}]

    if itype == "MAPPLET":
        mapplet_node = repo.get_mapplet(folder_name, tname)
        if mapplet_node is None:
            return [{"path": path, "status": "deadend"}]
        inner_ctx = make_lineage_ctx(mapplet_node, parent_ctx=ctx, parent_instance_name=instance_name)
        input_instance = _lineage_find_boundary_instance(inner_ctx, "input")
        if input_instance is None:
            return [{"path": path, "status": "deadend"}]
        return trace_lineage(inner_ctx, folder_name, repo, input_instance, field_name,
                              visited, path, depth + 1, max_depth)

    t_node = _lineage_transformation_node(ctx, folder_name, repo, tname) if itype == "TRANSFORMATION" else None

    # Bridge back out of a mapplet once we reach its Output boundary transformation
    if t_node is not None and "output" in attr(t_node, "TYPE", "").lower() and ctx["parent_ctx"] is not None:
        parent_ctx = ctx["parent_ctx"]
        nexts = parent_ctx["outgoing"].get((ctx["parent_instance_name"], field_name), [])
        if not nexts:
            return [{"path": path, "status": "deadend"}]
        results = []
        for ni, nf in nexts:
            results += trace_lineage(parent_ctx, folder_name, repo, ni, nf, visited, path, depth + 1, max_depth)
        return results

    output_fields = _lineage_resolve_outputs(t_node, field_name) if t_node is not None else [field_name]
    if not output_fields:
        fallback = _lineage_unconnected_fallback(
            ctx, folder_name, repo, t_node, tname, visited, path, depth, max_depth)
        if fallback is not None:
            return fallback
        return [{"path": path, "status": "deadend"}]

    results = []
    for of in output_fields:
        sub_path = path if of == field_name else path[:-1] + [f"{instance_name}.{of}"]
        nexts = ctx["outgoing"].get((instance_name, of), [])
        if not nexts:
            fallback = _lineage_unconnected_fallback(
                ctx, folder_name, repo, t_node, tname, visited, sub_path, depth, max_depth)
            if fallback is not None:
                results += fallback
            else:
                results.append({"path": sub_path, "status": "deadend"})
        else:
            for ni, nf in nexts:
                results += trace_lineage(ctx, folder_name, repo, ni, nf, visited, sub_path, depth + 1, max_depth)
    return results


def find_mapping_anywhere(repo, mapping_name, case_sensitive):
    """Look up a mapping by name across every folder (first match wins)."""
    for folder in repo.folders:
        fname = attr(folder, "NAME")
        for name, node in repo.folder_mappings.get(fname, {}).items():
            if names_equal(name, mapping_name, case_sensitive):
                return fname, node
    return None, None


LINEAGE_COLUMNS = [
    "Mapping_name", "Transformation_name", "Transformation_type",
    "Transformation_Field", "Lineage", "Status",
]
INVOLVED = "Involved Reporting Lineage"
NOT_INVOLVED = "Not_Involved"


def _trace_occurrence_lineage(ctx_for_mapping, folder_name, repo, t_full_name,
                               field_value, attribute):
    """Trace one occurrence (as returned by search_mapping_with_mapplets) forward
    through its own mapping's connector graph, crossing into/out of mapplets as
    needed. Returns (lineage_str, reached_sink_names) where reached_sink_names is
    the set of every TARGET tname/instance name any traced path actually landed
    on (empty if every path dead-ended)."""
    tokens = [tok.strip() for tok in field_value.split(";")
              if tok.strip() and not tok.strip().startswith("[")]
    if not tokens:
        tokens = [attribute]

    if "::" in t_full_name:
        mapplet_name, raw_instance = t_full_name.split("::", 1)
        mapplet_node = repo.get_mapplet(folder_name, mapplet_name)
        parent_instance_name = None
        for inst_name, inst in ctx_for_mapping["inst_by_name"].items():
            if attr(inst, "TYPE", "").upper() == "MAPPLET" and \
                    names_equal(attr(inst, "TRANSFORMATION_NAME"), mapplet_name, False):
                parent_instance_name = inst_name
                break
        if mapplet_node is None or parent_instance_name is None:
            start_ctx, start_instance = None, raw_instance
        else:
            start_ctx = make_lineage_ctx(mapplet_node, parent_ctx=ctx_for_mapping,
                                          parent_instance_name=parent_instance_name)
            start_instance = raw_instance
    else:
        start_ctx, start_instance = ctx_for_mapping, t_full_name

    terms = []
    if start_ctx is not None:
        for tok in tokens:
            terms.extend(trace_lineage(start_ctx, folder_name, repo, start_instance, tok, set(), []))
    else:
        terms = [{"path": [f"{t_full_name}.{tokens[0]}"], "status": "deadend"}]

    path_strs = []
    reached_sinks = set()
    for term in terms:
        path_strs.append(" -> ".join(term["path"]))
        if term["status"] == "sink":
            sink_tname = term.get("sink_tname", "")
            sink_inst = term.get("sink_instance", "")
            if sink_tname:
                reached_sinks.add(sink_tname)
            if sink_inst:
                reached_sinks.add(sink_inst)

    lineage_str = " | ".join(dict.fromkeys(path_strs))  # de-dup, keep order
    return lineage_str, reached_sinks


def build_reporting_lineage_rows(root, attribute, case_sensitive, expand_mapplets,
                                  target_mapping, target_table):
    """
    Tab 2: one row for EVERY occurrence of `attribute` found anywhere in the
    repository - the exact same set of occurrences as Tab 1 (same walk:
    every WORKFLOW/SESSION/MAPPING, plus mappings not wired to any session).
    Each occurrence is traced forward through the connector graph of its OWN
    mapping (crossing into/out of mapplets, following derived ports), so
    every row always shows a lineage path.

    A row is only marked INVOLVED when both:
      * its occurrence's mapping matches `target_mapping` by name, AND
      * the traced lineage actually reaches an instance/target pointing at
        `target_table` inside that mapping.
    Every other occurrence is still included in the tab, marked
    NOT_INVOLVED, with its own lineage traced within its own mapping.
    """
    repo = Repository(root)
    mapping_cache = {}
    lineage_ctx_cache = {}
    rows = []
    mappings_seen_via_session = set()

    def get_ctx(mapping):
        key = id(mapping)
        ctx = lineage_ctx_cache.get(key)
        if ctx is None:
            ctx = make_lineage_ctx(mapping)
            lineage_ctx_cache[key] = ctx
        return ctx

    def process(folder_name, mapping_name, mapping):
        matches = search_mapping_with_mapplets(
            repo, folder_name, mapping, attribute, case_sensitive, expand_mapplets, mapping_cache)
        if not matches:
            return
        ctx = get_ctx(mapping)
        is_target_mapping = names_equal(mapping_name, target_mapping, False)
        for m in matches:
            t_full_name = m["transformation_name"]
            t_type = m["transformation_type"]
            field_value = m["field"]

            lineage_str, reached_sinks = _trace_occurrence_lineage(
                ctx, folder_name, repo, t_full_name, field_value, attribute)

            reached_target = is_target_mapping and any(
                names_equal(s, target_table, False) for s in reached_sinks)

            rows.append({
                "Mapping_name": mapping_name,
                "Transformation_name": t_full_name,
                "Transformation_type": t_type,
                "Transformation_Field": field_value,
                "Lineage": lineage_str,
                "Status": INVOLVED if reached_target else NOT_INVOLVED,
            })

    for folder in repo.folders:
        folder_name = attr(folder, "NAME")

        for wf in children(folder, "WORKFLOW"):
            for sess in children(wf, "SESSION"):
                mapping_name = attr(sess, "MAPPINGNAME")
                mapping = repo.get_mapping(folder_name, mapping_name)
                if mapping is None:
                    continue
                mappings_seen_via_session.add(id(mapping))
                process(folder_name, mapping_name, mapping)

        # mappings that exist but are not (yet) wired to any workflow/session
        for mapping_name, mapping in repo.folder_mappings.get(folder_name, {}).items():
            if id(mapping) in mappings_seen_via_session:
                continue
            process(folder_name, mapping_name, mapping)

    return rows


# ---------------------------------------------------------------------------
# 6. Top-level walk: FOLDER -> WORKFLOW -> SESSION -> MAPPING
# ---------------------------------------------------------------------------

def find_occurrences(root, attribute, case_sensitive=False, expand_mapplets=True):
    repo = Repository(root)
    mapping_cache = {}
    rows = []
    mappings_seen_via_session = set()

    for folder in repo.folders:
        folder_name = attr(folder, "NAME")

        for wf in children(folder, "WORKFLOW"):
            wf_name = attr(wf, "NAME")

            for sess in children(wf, "SESSION"):
                sess_name = attr(sess, "NAME")
                mapping_name = attr(sess, "MAPPINGNAME")
                mapping = repo.get_mapping(folder_name, mapping_name)
                if mapping is None:
                    continue
                mappings_seen_via_session.add(id(mapping))

                matches = search_mapping_with_mapplets(
                    repo, folder_name, mapping, attribute,
                    case_sensitive, expand_mapplets, mapping_cache)

                for m in matches:
                    rows.append({
                        "Workflow_name": wf_name,
                        "Session_name": sess_name,
                        "Mapping_name": mapping_name,
                        "Transformation_name": m["transformation_name"],
                        "Transformation_type": m["transformation_type"],
                        "Transformation_Field": m["field"],
                        "Business_Logic": m["business_logic"],
                        "Category": m["category"],
                        "Predecessor_Input": m["predecessor"],
                        "Successor_Output": m["successor"],
                        "Additional_Information": m.get("additional_info", ""),
                    })

        # mappings that exist but are not (yet) wired to any workflow/session
        for mapping_name, mapping in repo.folder_mappings.get(folder_name, {}).items():
            if id(mapping) in mappings_seen_via_session:
                continue
            matches = search_mapping_with_mapplets(
                repo, folder_name, mapping, attribute,
                case_sensitive, expand_mapplets, mapping_cache)
            for m in matches:
                rows.append({
                    "Workflow_name": "(not linked to a Workflow)",
                    "Session_name": "(not linked to a Session)",
                    "Mapping_name": mapping_name,
                    "Transformation_name": m["transformation_name"],
                    "Transformation_type": m["transformation_type"],
                    "Transformation_Field": m["field"],
                    "Business_Logic": m["business_logic"],
                    "Category": m["category"],
                    "Predecessor_Input": m["predecessor"],
                    "Successor_Output": m["successor"],
                    "Additional_Information": m.get("additional_info", ""),
                })

    return rows


# ---------------------------------------------------------------------------
# 7. Excel output
# ---------------------------------------------------------------------------

COLUMNS = [
    "Workflow_name", "Session_name", "Mapping_name", "Transformation_name",
    "Transformation_type", "Transformation_Field", "Business_Logic",
    "Category", "Predecessor_Input", "Successor_Output", "Additional_Information",
]


def write_workbook(rows, attribute, out_path,
                    lineage_rows=None, lineage_mapping_name=None, lineage_target_table=None):
    wb = Workbook()

    # ---- Tab 1: Occurrences -----------------------------------------
    ws = wb.active
    ws.title = "Occurrences"

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    body_font = Font(name="Arial", size=10)
    direct_fill = PatternFill("solid", fgColor="E2EFDA")
    indirect_fill = PatternFill("solid", fgColor="FFF2CC")

    ws.append(COLUMNS)
    for c in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"

    MAX_CELL_LEN = 32000  # stay under Excel's 32767 hard limit

    for r in rows:
        values = []
        for col in COLUMNS:
            v = r[col]
            if isinstance(v, str) and len(v) > MAX_CELL_LEN:
                v = v[:MAX_CELL_LEN] + " ...[truncated]"
            values.append(v)
        ws.append(values)
        row_idx = ws.max_row
        fill = direct_fill if r["Category"] == DIRECT else indirect_fill
        for c in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.font = body_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if c == COLUMNS.index("Category") + 1:
                cell.fill = fill

    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{ws.max_row}"

    widths = [22, 22, 22, 26, 20, 24, 46, 20, 30, 30, 44]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---- Tab 2: Reporting_Lineage --------------------------------------
    if lineage_rows is not None:
        wsl = wb.create_sheet("Reporting_Lineage", 1)
        wsl.append(LINEAGE_COLUMNS)
        for c in range(1, len(LINEAGE_COLUMNS) + 1):
            cell = wsl.cell(row=1, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        wsl.freeze_panes = "A2"

        involved_fill = PatternFill("solid", fgColor="C6E0B4")
        not_involved_fill = PatternFill("solid", fgColor="F8CBAD")

        for r in lineage_rows:
            values = []
            for col in LINEAGE_COLUMNS:
                v = r[col]
                if isinstance(v, str) and len(v) > MAX_CELL_LEN:
                    v = v[:MAX_CELL_LEN] + " ...[truncated]"
                values.append(v)
            wsl.append(values)
            row_idx = wsl.max_row
            fill = involved_fill if r["Status"] == INVOLVED else not_involved_fill
            for c in range(1, len(LINEAGE_COLUMNS) + 1):
                cell = wsl.cell(row=row_idx, column=c)
                cell.font = body_font
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if c == LINEAGE_COLUMNS.index("Status") + 1:
                    cell.fill = fill

        if lineage_rows:
            wsl.auto_filter.ref = f"A1:{get_column_letter(len(LINEAGE_COLUMNS))}{wsl.max_row}"

        widths_l = [22, 26, 20, 24, 80, 24]
        for i, w in enumerate(widths_l, start=1):
            wsl.column_dimensions[get_column_letter(i)].width = w

    # ---- Tab 3: Summary ----------------------------------------------
    ws2 = wb.create_sheet("Summary")
    ws2.append(["Attribute searched", attribute])
    ws2.append(["Total occurrences", len(rows)])
    ws2.append(["Direct Pass through occurrences", sum(1 for r in rows if r["Category"] == DIRECT)])
    ws2.append(["Involves Derivation occurrences", sum(1 for r in rows if r["Category"] == INDIRECT)])
    ws2.append(["Distinct workflows", len({r["Workflow_name"] for r in rows})])
    ws2.append(["Distinct sessions", len({r["Session_name"] for r in rows})])
    ws2.append(["Distinct mappings", len({r["Mapping_name"] for r in rows})])
    ws2.append(["Distinct transformations",
                len({(r["Mapping_name"], r["Transformation_name"]) for r in rows})])
    for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, max_col=2):
        row[0].font = Font(name="Arial", bold=True)
        row[1].font = Font(name="Arial")
    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 40

    # breakdown by mapping
    ws2.append([])
    ws2.append(["Mapping_name", "Occurrences"])
    ws2.cell(row=ws2.max_row, column=1).font = Font(name="Arial", bold=True)
    ws2.cell(row=ws2.max_row, column=2).font = Font(name="Arial", bold=True)
    counts = defaultdict(int)
    for r in rows:
        counts[r["Mapping_name"]] += 1
    for mapping_name, n in sorted(counts.items(), key=lambda kv: -kv[1]):
        ws2.append([mapping_name, n])

    if lineage_rows is not None:
        ws2.append([])
        ws2.append(["Reporting Lineage - Mapping", lineage_mapping_name])
        ws2.append(["Reporting Lineage - Target table", lineage_target_table])
        ws2.append(["Reporting Lineage - Occurrences traced", len(lineage_rows)])
        ws2.append(["Reporting Lineage - Involved Reporting Lineage",
                    sum(1 for r in lineage_rows if r["Status"] == INVOLVED)])
        ws2.append(["Reporting Lineage - Not_Involved",
                    sum(1 for r in lineage_rows if r["Status"] == NOT_INVOLVED)])
        for row in ws2.iter_rows(min_row=ws2.max_row - 4, max_row=ws2.max_row, max_col=2):
            row[0].font = Font(name="Arial", bold=True)
            row[1].font = Font(name="Arial")

    wb.save(out_path)


# ---------------------------------------------------------------------------
# 8. CLI
# ---------------------------------------------------------------------------

def main():
    p = argparse.ArgumentParser(
        description="Find every occurrence of an attribute/field name across "
                    "an Informatica PowerCenter XML/JSON export and export "
                    "the lineage to Excel.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__)
    p.add_argument("--input", "-i", required=True, help="Path to the .xml or .json export")
    p.add_argument("--attribute", "-a", required=False,
                   help="Attribute/field name to search for. If omitted, you'll be prompted.")
    p.add_argument("--output", "-o", default=None,
                   help="Output .xlsx path (default: <attribute>_lineage.xlsx)")
    p.add_argument("--case-sensitive", action="store_true",
                   help="Match the attribute name case-sensitively (default: case-insensitive)")
    p.add_argument("--no-mapplets", action="store_true",
                   help="Do not expand/search inside mapplets referenced by mappings")
    p.add_argument("--target-mapping", "-tm", default=None,
                   help="Reporting Mapping name to build the Reporting_Lineage tab (Tab 2) for. "
                        "Requires --target-table. If neither is given, you'll be prompted "
                        "(leave blank to skip Tab 2).")
    p.add_argument("--target-table", "-tt", default=None,
                   help="Reporting Target table/instance name that marks a 'reached the sink' "
                        "lineage in the Reporting_Lineage tab (Tab 2). Requires --target-mapping.")
    p.add_argument("--no-lineage-tab", action="store_true",
                   help="Skip the Reporting_Lineage tab entirely and don't prompt for it.")
    args = p.parse_args()

    attribute = args.attribute or input("Attribute name to search for: ").strip()
    if not attribute:
        sys.exit("No attribute name given.")

    root = load_tree(args.input)

    print(f"Searching for '{attribute}' "
          f"({'case-sensitive' if args.case_sensitive else 'case-insensitive'})...")
    rows = find_occurrences(
        root, attribute,
        case_sensitive=args.case_sensitive,
        expand_mapplets=not args.no_mapplets)

    target_mapping, target_table = args.target_mapping, args.target_table
    if not args.no_lineage_tab and target_mapping is None and target_table is None:
        tm = input("Reporting Mapping name for the Reporting_Lineage tab "
                   "(leave blank to skip Tab 2): ").strip()
        if tm:
            tt = input("Reporting Target table name (sink) for the Reporting_Lineage tab: ").strip()
            target_mapping, target_table = tm, tt or None

    lineage_rows = None
    if target_mapping and target_table:
        print(f"Building Reporting_Lineage tab for all occurrences of '{attribute}', "
              f"marking those in mapping '{target_mapping}' that reach target "
              f"'{target_table}' as Involved...")
        _, found_mapping = find_mapping_anywhere(Repository(root), target_mapping, args.case_sensitive)
        if found_mapping is None:
            print(f"Warning: mapping '{target_mapping}' was not found in the repository; "
                  f"every row in the Reporting_Lineage tab will be marked '{NOT_INVOLVED}'.")
        lineage_rows = build_reporting_lineage_rows(
            root, attribute, case_sensitive=args.case_sensitive,
            expand_mapplets=not args.no_mapplets,
            target_mapping=target_mapping, target_table=target_table)
    elif target_mapping or target_table:
        print("Warning: both --target-mapping and --target-table are needed to build the "
              "Reporting_Lineage tab; skipping it.")

    out_path = args.output or f"{re.sub(r'[^A-Za-z0-9_-]+', '_', attribute)}_lineage.xlsx"
    write_workbook(rows, attribute, out_path,
                   lineage_rows=lineage_rows,
                   lineage_mapping_name=target_mapping,
                   lineage_target_table=target_table)

    print(f"Found {len(rows)} occurrence(s).")
    if lineage_rows is not None:
        involved = sum(1 for r in lineage_rows if r["Status"] == INVOLVED)
        print(f"Reporting lineage: {involved}/{len(lineage_rows)} occurrence(s) reach "
              f"'{target_table}' in '{target_mapping}'.")
    print(f"Workbook written to: {out_path}")


if __name__ == "__main__":
    main()
